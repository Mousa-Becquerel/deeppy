"""
Parser for user-uploaded Bill-of-Materials XLSX files.

Reads the BoM and returns a compact JSON-friendly structure that can be:
  1. Injected into the extraction prompt as authoritative context for AI
  2. Used directly to populate the composition section after extraction

Expected column structure (matches DeePPy_data_ontology.xlsx → Bill_of_materials):
  - Bill of materials (material id, e.g. "Material#1")
  - ID code
  - Description
  - Unit
  - Quantity (per product)
  - Supplier 1 - Name / Address / Means of transportation / EU Vehicle class / KM
  - Supplier 2 - Name / Address / Means of transportation / EU Vehicle class / KM
  - ... (variable number of suppliers)

The parser is forgiving: it auto-detects the header row and tolerates
column-name variations (it/en, abbreviations, extra whitespace).
"""

import logging
import re
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)


# ─── Column-name normalisation ────────────────────────────────────────────

def _norm(s: Any) -> str:
    """Lower-case + strip + collapse whitespace for fuzzy header matching."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).lower()).strip()


# Canonical column → list of accepted header substrings
_COL_PATTERNS: dict[str, list[str]] = {
    "material_id": ["bill of materials", "material id", "id materiale", "material#"],
    "id_code": ["id code", "codice", "sku"],
    "description": ["description", "descrizione", "materiale", "material"],
    "unit": ["unit", "unità", "udm", "u.m."],
    "quantity": ["quantity", "quantità", "qty", "qta", "qta'"],
}

_SUPPLIER_FIELD_PATTERNS: dict[str, list[str]] = {
    "name": ["name", "nome", "ragione sociale"],
    "address": ["address", "indirizzo"],
    # Tolerant of common typos: "trasportation", "trasporto"
    "transport_method": ["transport", "trasport", "means of", "mezzo"],
    # Tolerant of "Veihcle" (typo in ontology) and various EURO class wordings
    "eu_vehicle_class": ["vehicle class", "veihcle class", "veicolo", "classe euro", "eu class", "euro class"],
    "distance_km": [" km", "(km)", "distance", "distanza"],
}


def _match_pattern(header: str, patterns: list[str]) -> bool:
    h = _norm(header)
    return any(p in h for p in patterns)


def _detect_supplier_index(header: str) -> int | None:
    """Return supplier number (1, 2, 3, ...) if header looks like a supplier column."""
    h = _norm(header)
    m = re.search(r"supplier\s*(\d+)|fornitore\s*(\d+)", h)
    if m:
        return int(m.group(1) or m.group(2))
    return None


# ─── Sheet → row dict mapping ─────────────────────────────────────────────

def _build_column_map(header_row: list[Any]) -> dict[int, tuple[str, int | None]]:
    """
    Map column index → (canonical_field, supplier_index_or_None).

    e.g. col 5 might be ('name', 1) for "Supplier 1 - Name".
    """
    col_map: dict[int, tuple[str, int | None]] = {}
    for idx, val in enumerate(header_row):
        if val is None:
            continue
        sup_idx = _detect_supplier_index(str(val))
        for canonical, patterns in _COL_PATTERNS.items():
            if _match_pattern(val, patterns) and sup_idx is None:
                col_map[idx] = (canonical, None)
                break
        else:
            # Try supplier sub-fields
            if sup_idx is not None:
                for sub_field, patterns in _SUPPLIER_FIELD_PATTERNS.items():
                    if _match_pattern(val, patterns):
                        col_map[idx] = (sub_field, sup_idx)
                        break
    return col_map


def _find_header_row(ws) -> int | None:
    """
    Find the row that looks like a header.

    Strategy: scan first ~10 rows; the row whose cells most often match
    known patterns wins.
    """
    best_row = None
    best_score = 0
    for row_idx in range(1, min(ws.max_row + 1, 11)):
        row = [c.value for c in ws[row_idx]]
        score = 0
        for val in row:
            if val is None:
                continue
            for patterns in _COL_PATTERNS.values():
                if _match_pattern(val, patterns):
                    score += 1
                    break
            else:
                if _detect_supplier_index(str(val)) is not None:
                    score += 1
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row if best_score >= 2 else None


# ─── Public API ───────────────────────────────────────────────────────────

def parse_bom_xlsx(xlsx_path: Path) -> dict | None:
    """
    Parse a BoM XLSX file into a structured dict.

    Returns None if no recognisable BoM structure is found.

    Output shape:
      {
        "source_file": "EUROFINESTRA_VERSATILE_BoM.xlsx",
        "materials": [
          {
            "material_id": "Material#1",
            "id_code": "...",
            "description": "Wooden frame",
            "unit": "kg",
            "quantity": 12.4,
            "suppliers": [
              {"name": "...", "address": "...", "transport_method": "...",
               "eu_vehicle_class": "...", "distance_km": 120.0},
              ...
            ],
          },
          ...
        ],
      }
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        logger.warning(f"BoM parse failed for {xlsx_path.name}: {e}")
        return None

    # Try each sheet — pick the one that yields the most materials
    best: dict | None = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2 or ws.max_column < 2:
            continue
        result = _parse_sheet(ws)
        if result and (best is None or len(result["materials"]) > len(best["materials"])):
            best = result

    if best is None:
        logger.warning(f"No recognisable BoM table in {xlsx_path.name}")
        return None

    best["source_file"] = xlsx_path.name
    logger.info(
        f"Parsed BoM {xlsx_path.name}: {len(best['materials'])} materials, "
        f"{sum(len(m.get('suppliers', [])) for m in best['materials'])} supplier entries"
    )
    return best


def _parse_sheet(ws) -> dict | None:
    """Try to parse one sheet. Returns dict or None."""
    header_row_idx = _find_header_row(ws)
    if header_row_idx is None:
        return None

    header_row = [c.value for c in ws[header_row_idx]]
    col_map = _build_column_map(header_row)

    if "description" not in {c[0] for c in col_map.values()}:
        # No description column — not a real BoM
        return None

    materials = []
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row = [c.value for c in ws[row_idx]]
        # Skip fully empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        material: dict[str, Any] = {"suppliers": {}}
        for col_idx, (field, sup_idx) in col_map.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None or (isinstance(val, str) and not val.strip()):
                continue
            if sup_idx is None:
                # Top-level field
                material[field] = val
            else:
                # Supplier sub-field
                supplier = material["suppliers"].setdefault(sup_idx, {})
                if field == "distance_km":
                    try:
                        supplier[field] = float(val)
                    except (ValueError, TypeError):
                        supplier[field] = val
                else:
                    supplier[field] = val

        # Convert suppliers dict → ordered list
        material["suppliers"] = [material["suppliers"][i] for i in sorted(material["suppliers"])]

        # Skip rows without a description (placeholder rows)
        if not material.get("description"):
            continue

        # Coerce quantity to float
        if "quantity" in material:
            try:
                material["quantity"] = float(material["quantity"])
            except (ValueError, TypeError):
                pass

        materials.append(material)

    if not materials:
        return None
    return {"materials": materials}


def bom_to_prompt_snippet(bom: dict, max_chars: int = 4000) -> str:
    """
    Render a parsed BoM into a compact text block for the extraction prompt.

    Capped at max_chars to avoid blowing up the prompt.
    """
    if not bom or not bom.get("materials"):
        return ""

    lines = [
        "",
        f"USER-PROVIDED BILL OF MATERIALS (file: {bom.get('source_file', 'BoM.xlsx')}):",
        "This BoM is the SINGLE SOURCE OF TRUTH for composition. The "
        "composition.materials[] list you return must contain EXACTLY one entry "
        "per BoM row, with these fields populated from the BoM:",
        "  • description ← the material description from the BoM",
        "  • quantity ← the numeric value from the BoM 'qty=' (parse the number, drop units)",
        "  • unit ← the unit string from the BoM (e.g. 'kg', '%', 'pieces')",
        "  • material_id ← the 'id=' code from the BoM",
        "  • suppliers ← include name, address, transport_method, distance_km from the BoM",
        "DO NOT leave quantity/unit/id/suppliers null when the BoM has values.",
        "STRICT EXCLUSION RULES (read carefully — these are the most common errors):",
        "  ✗ DO NOT add 'duplicate' materials inferred from marketing text in the "
        "PDFs (e.g. when the tech sheet mentions 'made with sustainable fibers, "
        "natural binders and recycled content' — these are NOT separate materials).",
        "  ✗ DO NOT split a BoM row into multiple sub-components.",
        "  ✗ DO NOT add generic materials like 'Recycled fibers', 'Cork', "
        "'Perlite', 'Additives' UNLESS they appear as a labeled composition row "
        "in a formal EPD/DoP table with their own quantity.",
        "  ✓ You MAY add a material ONLY if a PDF has an EXPLICIT composition "
        "table row stating that material as a percentage or mass — never from prose.",
        "  ✓ Default: same number of materials as BoM rows. Adding extras is the "
        "exception, requires a clear formal table source.",
        "",
    ]
    for i, m in enumerate(bom["materials"], 1):
        parts = [f"{i}. {m.get('description', 'Unknown')}"]
        if m.get("id_code"):
            parts.append(f"id={m['id_code']}")
        if m.get("quantity") is not None:
            parts.append(f"qty={m['quantity']}{(' ' + m['unit']) if m.get('unit') else ''}")
        line = "  " + " · ".join(parts)
        if m.get("suppliers"):
            for j, s in enumerate(m["suppliers"], 1):
                supplier_parts = []
                if s.get("name"):
                    supplier_parts.append(s["name"])
                if s.get("address"):
                    supplier_parts.append(s["address"])
                if s.get("transport_method"):
                    supplier_parts.append(s["transport_method"])
                if s.get("eu_vehicle_class"):
                    supplier_parts.append(s["eu_vehicle_class"])
                if s.get("distance_km") is not None:
                    supplier_parts.append(f"{s['distance_km']} km")
                if supplier_parts:
                    line += f"\n     Supplier {j}: {' / '.join(str(p) for p in supplier_parts)}"
        lines.append(line)

    text = "\n".join(lines)
    if len(text) > max_chars:
        text = text[:max_chars] + "\n... [truncated]"
    return text
