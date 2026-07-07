"""
Recall vs ground truth — compare an extracted passport against an expert
"expected results" spec (the ontology-style XLSX with a 'Dato output' column).

Recall = (expected fields the passport actually filled) / (expected fields).
This is the production version of the manual comparison: the user attaches the
expected-results file to a product and we score against it.
"""

from __future__ import annotations

import re
from typing import Optional

import openpyxl


# ── Parse the expected-results XLSX ──────────────────────────────────────────

def parse_expected_xlsx(path) -> list[dict]:
    """Read the expected-results sheet → rows that SHOULD have a value.

    Returns [{section, tier2, tier3, value}], skipping headers and rows whose
    'Dato output' is empty or '-'.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict] = []
    for r in ws.iter_rows(values_only=True):
        c = list(r) + [None] * 11
        section = str(c[0]).strip() if c[0] else ""
        if not section or section.lower().startswith("deeppy section"):
            continue
        tier2 = str(c[2] or "").strip()
        tier3 = str(c[3] or "").strip()
        dato = c[9]  # 'Dato output' column
        val = str(dato).strip() if dato is not None else ""
        if val in ("", "-"):
            continue
        rows.append({"section": section, "tier2": tier2, "tier3": tier3, "value": val})
    return rows


# ── Helpers ──────────────────────────────────────────────────────────────────

def _norm(s) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()


def _ev(o, *path):
    cur = o or {}
    for k in path:
        if not isinstance(cur, dict):
            return None
        cur = cur.get(k)
        if cur is None:
            return None
    return cur.get("value") if isinstance(cur, dict) and "value" in cur else cur


def _filled(v) -> bool:
    return v not in (None, "", [], {})


# Tier3-label (normalized) → passport path
_OVERVIEW = {
    "product name": ("overview", "product_info", "product_name"),
    "product image": ("overview", "product_info", "product_image"),
    "product description": ("overview", "product_info", "product_description"),
    "item type": ("overview", "product_info", "item_type"),
    "product family": ("overview", "product_info", "product_family"),
    "intended use": ("overview", "product_info", "intended_use"),
    "functional unit": ("overview", "product_info", "functional_unit"),
    "standard dimension wxdxh": ("overview", "product_info", "standard_dimension"),
    "standard dimension": ("overview", "product_info", "standard_dimension"),
    "weight": ("overview", "product_info", "weight"),
    "production period": ("overview", "product_info", "production_period"),
    "company name": ("overview", "manufacturer", "company_name"),
    "company description": ("overview", "manufacturer", "company_description"),
    "address": ("overview", "manufacturer", "address"),
    "website": ("overview", "manufacturer", "website"),
    "manufacturing site": ("overview", "manufacturer", "manufacturing_site"),
    "mail": ("overview", "manufacturer", "email"),
    "phone": ("overview", "manufacturer", "phone"),
    "sale type": ("overview", "manufacturer", "sale_type"),
}

_SAFETY = {
    "cmrs": "contains_cmrs", "svhcs": "contains_svhcs", "pentane": "contains_pentane",
    "pfas": "contains_pfas", "flame retardancy": "has_flame_retardancy",
    "rohs": "complies_rohs", "voc emission": "produces_voc",
    "heavy metals": "contains_heavy_metals", "asbestos": "contains_asbestos",
    "child labor": "complies_child_labor",
}

_DOCSLOT = {
    "2d model": "drawing_2d", "3d model": "drawing_3d",
    "method statement for installation": "method_statement_installation",
    "method statement for maintenance and repair": "method_statement_maintenance",
    "method statement for replacement refurbishment": "method_statement_replacement",
    "method statement for dismantling eol": "method_statement_dismantling",
    "technical data sheet": "technical_data_sheet",
    "brochure and commercial presentation": "brochure",
    "other documents": "other_documents",
}

_PERF_STOP = {"the", "of", "to", "and", "class", "avg", "resistance", "thermal", "trasmittance", "transmittance"}


def _levenshtein_le(a: str, b: str, k: int = 2) -> bool:
    """Return True if edit distance between a and b is <= k. O(len(a)*k) early-exit.

    Used to absorb minor typos in expert ground-truth labels
    (e.g. 'conducivity' vs 'conductivity', 'strenght' vs 'strength')."""
    if a == b:
        return True
    la, lb = len(a), len(b)
    if abs(la - lb) > k:
        return False
    prev = list(range(lb + 1))
    for i, ca in enumerate(a, 1):
        cur = [i] + [0] * lb
        min_row = cur[0]
        for j, cb in enumerate(b, 1):
            cost = 0 if ca == cb else 1
            cur[j] = min(cur[j - 1] + 1, prev[j] + 1, prev[j - 1] + cost)
            if cur[j] < min_row:
                min_row = cur[j]
        if min_row > k:
            return False
        prev = cur
    return prev[lb] <= k


def _fuzzy_token_in(tok: str, names: list[str]) -> bool:
    """True if `tok` appears as a substring in any name OR matches a token in
    any name with edit distance <= 2 (catches typos like 'conducivity'/'conductivity')."""
    if not tok:
        return False
    for n in names:
        if tok in n:
            return True
        for nt in n.split():
            if len(nt) >= 4 and _levenshtein_le(tok, nt, 2):
                return True
    return False


def _check_row(pp: dict, row: dict, mats: list[str], perf: list[str], docs: dict) -> bool:
    sec = row["section"].lower()
    t3n = _norm(row["tier3"])
    t2n = _norm(row["tier2"])

    if sec.startswith("overview"):
        path = None
        for key, p in _OVERVIEW.items():
            if key and (key in t3n or t3n == key):
                path = p
                break
        if not path:
            return False
        v = _ev(pp, *path)
        if path[-1] == "product_family" and not _filled(v):
            v = _ev(pp, "overview", "product_info", "product_family_code")
        return _filled(v)

    if sec.startswith("composition"):
        ev = _norm(row["value"])
        kw = ev.split()[0] if ev else ""
        return any(kw and kw in m for m in mats) if kw else len(mats) > 0

    if sec.startswith("performance"):
        label = _norm(re.sub(r"\[.*?\]", "", row["tier3"]))
        # Keep parenthetical codes like (uw)/(ug)/(rw) — they're discriminating.
        toks = [t for t in label.split() if t not in _PERF_STOP and len(t) > 1]
        if not toks:
            return False
        # match if the two most-specific tokens both appear in a perf name
        for pn in perf:
            if all(t in pn for t in toks[:2]):
                return True
        # fallback: the single most-specific token (substring or fuzzy)
        return _fuzzy_token_in(toks[0], perf)

    if sec.startswith("compliance"):
        c = pp.get("compliance") or {}
        if t3n == "dop":
            return _filled(_ev(c, "dop_reference"))
        if t3n == "doc":
            return _filled(_ev(c, "doc_reference"))
        if t3n == "ce":
            return _filled(_ev(c, "ce_marking"))
        if t3n == "quality":
            return _filled(_ev(c, "quality_control"))
        if "certif" in t3n:
            return len(c.get("product_certifications") or []) > 0
        if "other labels" in t3n or "other labels" in t2n:
            return _filled(_ev(c, "other_labels"))
        sf = c.get("safety") or {}
        for kw, field in _SAFETY.items():
            if kw in t3n:
                return _filled(_ev(sf, field))
        return False

    if sec.startswith("documents"):
        for kw, slot in _DOCSLOT.items():
            if kw in t3n or (t2n and kw in t2n):
                return bool(docs.get(slot))
        return False

    return False


def compute_recall(passport: dict, rows: list[dict]) -> dict:
    """Score the passport against the expected rows. Returns recall + misses."""
    pp = passport or {}
    mats = [_norm(_ev(m, "description")) for m in (pp.get("composition") or {}).get("materials") or []]
    perf = [_norm(_ev(v, "property_name")) for v in (pp.get("performance") or {}).get("values") or []]
    docs = pp.get("documents") or {}

    captured = 0
    missed: list[str] = []
    for row in rows:
        if _check_row(pp, row, mats, perf, docs):
            captured += 1
        else:
            missed.append(row["tier3"] or row["tier2"] or row["section"])

    total = len(rows)
    return {
        "recall": round(captured / total * 100, 1) if total else 0.0,
        "captured": captured,
        "total": total,
        "missed": missed,
    }
